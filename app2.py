# =====================================
# Streamlit Web App: 模拟Project：人事用合同记录表自动审核（四输出表版 + 漏填检查 + 驻店客户版）
# =====================================

import streamlit as st
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from io import BytesIO

def normalize_contract_key(series: pd.Series) -> pd.Series:
    """
    对合同号 Series 进行标准化处理，用于安全的 pd.merge 操作。
    """
    # 1. 确保是字符串类型，同时处理缺失值（如果存在）
    s = series.astype(str)
    
    # 2. 移除常见的浮点数残留（以防原始数据错误输入）
    s = s.str.replace(r"\.0$", "", regex=True) 
    
    # 3. 核心：移除首尾空格（处理最常见的导入错误）
    s = s.str.strip()
    
    # 4. 统一转换为大写（处理大小写不一致问题，如 'pazl' vs 'PAZL'）
    s = s.str.upper() 
    
    # 5. 处理全角/半角差异（将常见的全角连接符转为半角）
    s = s.str.replace('－', '-', regex=False) # 全角连接符转半角
    
    # 6. 处理其他可能的空白字符（例如 tabs, 换行符等）
    s = s.str.replace(r'\s+', '', regex=True)
    
    return s

# =====================================
# 🏁 应用标题与说明
# =====================================
st.title("📊 模拟实际运用环境Project：人事用合同记录表自动审核系统（四Sheet + 漏填检查 + 驻店客户版）")

# =====================================
# 📂 上传文件区：要求上传 5 个 xlsx 文件
# =====================================
uploaded_files = st.file_uploader(
    "请上传以下文件：记录表、放款明细、字段、二次明细、重卡数据",
    type="xlsx",
    accept_multiple_files=True
)

if not uploaded_files or len(uploaded_files) < 5:
    st.warning("⚠️ 请上传所有 5 个文件后继续")
    st.stop()
else:
    st.success("✅ 文件上传完成")

# =====================================
# 🧰 工具函数区（文件定位、列名模糊匹配、日期/数值处理）
# =====================================

# 按关键字查找文件（文件名包含关键字）
def find_file(files_list, keyword):
    for f in files_list:
        if keyword in f.name:
            return f
    raise FileNotFoundError(f"❌ 未找到包含关键词「{keyword}」的文件")

# 统一列名格式（去空格、转小写）
def normalize_colname(c): return str(c).strip().lower()

# 按关键字匹配列名（支持 exact 精确匹配与模糊匹配）
def find_col(df, keyword, exact=False):
    key = keyword.strip().lower()
    for col in df.columns:
        cname = normalize_colname(col)
        if (exact and cname == key) or (not exact and key in cname):
            return col
    return None

# 查找sheet（sheet名包含关键字即可）
def find_sheet(xls, keyword):
    for s in xls.sheet_names:
        if keyword in s:
            return s
    raise ValueError(f"❌ 未找到包含关键词「{keyword}」的sheet")

# 统一数值解析（去逗号、转float、处理百分号）
def normalize_num(val):
    if pd.isna(val): return None
    s = str(val).replace(",", "").strip()
    if s in ["", "-", "nan"]: return None
    try:
        if "%" in s: return float(s.replace("%", "")) / 100
        return float(s)
    except ValueError:
        return s

# 日期匹配（年/月/日完全一致）
def same_date_ymd(a, b):
    try:
        da = pd.to_datetime(a, errors='coerce')
        db = pd.to_datetime(b, errors='coerce')
        if pd.isna(da) or pd.isna(db): return False
        return (da.year, da.month, da.day) == (db.year, db.month, db.day)
    except Exception:
        return False
def prepare_ref_df(ref_df, mapping, prefix):
    # 假设合同号列名为 contract_col
    ref_contract_col = [k for k, v in mapping.items() if v == '合同号'][0]
    
    std_df = pd.DataFrame()
    
    # VVVV 插入归一化函数 VVVV
    std_df['__KEY__'] = normalize_contract_key(ref_df[ref_contract_col])
    
    # 提取并重命名所有需要的字段
    for main_kw, ref_kw in mapping.items():
        # 城市经理需要精确匹配
        exact = (main_kw == "城市经理") 
        ref_col_name = find_col(ref_df, ref_kw, exact=exact)
        
        if ref_col_name:
            # 使用标准化的列名
            std_df[f'ref_{prefix}_{main_kw}'] = ref_df[ref_col_name]
        else:
            st.warning(f"⚠️ 在 {prefix} 参考表中未找到列 (main: '{main_kw}', ref: '{ref_kw}')")

    # 效仿原始逻辑：只取第一个匹配项
    std_df = std_df.drop_duplicates(subset=['__KEY__'], keep='first')
    return std_df

def compare_series_vec(s_main, s_ref, main_kw):
    """
    向量化比较两个Series，复刻原始的 compare_fields_and_mark 逻辑。
    返回一个布尔Series，True表示存在差异。
    """
    
    # 1. 预处理：处理空值。原始逻辑：同为NaN/空则认为一致。
    main_is_na = pd.isna(s_main) | (s_main.astype(str).str.strip().isin(["", "nan", "None"]))
    ref_is_na = pd.isna(s_ref) | (s_ref.astype(str).str.strip().isin(["", "nan", "None"]))
    
    # 如果两者都为空，则不算错误
    both_are_na = main_is_na & ref_is_na
    
    # 如果参考值为空，不应报错 (模仿 ref_rows.empty)
    # （merge后，未匹配到的行 ref_is_na 会为 True，这里我们只关心两者都为空的情况）

    # 2. 日期字段比较
    if any(k in main_kw for k in ["日期", "时间"]):
        d_main = pd.to_datetime(s_main, errors='coerce')
        d_ref = pd.to_datetime(s_ref, errors='coerce')
        
        # 原始 same_date_ymd 逻辑：
        # 只有当两者都是有效日期且年月日不相等时，才算错误。
        valid_dates_mask = d_main.notna() & d_ref.notna()
        date_diff_mask = (d_main.dt.date != d_ref.dt.date)
        
        errors = valid_dates_mask & date_diff_mask
    
    # 3. 数值/文本比较
    else:
        # 使用原始的 normalize_num 函数，但通过 apply 应用
        s_main_norm = s_main.apply(normalize_num)
        s_ref_norm = s_ref.apply(normalize_num)
        
        # 重新检查标准化后的空值
        main_is_na_norm = pd.isna(s_main_norm) | (s_main_norm.astype(str).str.strip().isin(["", "nan", "None"]))
        ref_is_na_norm = pd.isna(s_ref_norm) | (s_ref_norm.astype(str).str.strip().isin(["", "nan", "None"]))
        both_are_na_norm = main_is_na_norm & ref_is_na_norm

        # 检查是否为数值类型
        is_num_main = s_main_norm.apply(lambda x: isinstance(x, (int, float)))
        is_num_ref = s_ref_norm.apply(lambda x: isinstance(x, (int, float)))
        both_are_num = is_num_main & is_num_ref
        
        # 初始化错误Series
        errors = pd.Series(False, index=s_main.index)
        
        # 3a. 数值比较
        if both_are_num.any():
            num_main = s_main_norm[both_are_num]
            num_ref = s_ref_norm[both_are_num]
            diff = (num_main - num_ref).abs()
            
            if main_kw == "保证金比例":
                num_errors = (diff > 0.00500001) # 增加微小容差
            else:
                num_errors = (diff > 1e-6)
            
            errors.loc[both_are_num] = num_errors

        # 3b. 文本比较
        not_num_mask = ~both_are_num
        if not_num_mask.any():
            str_main = s_main_norm[not_num_mask].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            str_ref = s_ref_norm[not_num_mask].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            
            str_errors = (str_main != str_ref)
            errors.loc[not_num_mask] = str_errors
            
        # 最终错误：排除掉那些两者皆为空的情况
        errors = errors & ~both_are_na_norm
        return errors

    # 最终错误：排除掉那些两者皆为空的情况
    return errors & ~both_are_na

# =====================================
# 🧮 单sheet检查函数 (向量化版)
# =====================================
def check_one_sheet(sheet_keyword, main_file, ref_dfs_std_dict):
    start_time = time.time()
    xls_main = pd.ExcelFile(main_file)

    # 查找目标sheet
    try:
        target_sheet = find_sheet(xls_main, sheet_keyword)
    except ValueError:
        st.warning(f"⚠️ 未找到包含「{sheet_keyword}」的sheet，跳过。")
        return 0, None, 0, set()

    # 1. 读取目标sheet（第二行为表头）
    try:
        main_df = pd.read_excel(xls_main, sheet_name=target_sheet, header=1)
    except Exception as e:
        st.error(f"❌ 读取「{sheet_keyword}」时出错: {e}")
        return 0, None, 0, set()
        
    if main_df.empty:
        st.warning(f"⚠️ 「{sheet_keyword}」为空，跳过。")
        return 0, None, 0, set()

    # 2. 查找合同号列
    global contract_col_main
    contract_col_main = find_col(main_df, "合同")
    if not contract_col_main:
        st.error(f"❌ 在「{sheet_keyword}」中未找到合同列。")
        return 0, None, 0, set()

    # 3. 创建临时输出文件 (保留原始表头空行)
    output_path = f"记录表_{sheet_keyword}_审核标注版.xlsx"
    empty_row = pd.DataFrame([[""] * len(main_df.columns)], columns=main_df.columns)
    # 注意：这里我们保存的是原始main_df
    pd.concat([empty_row, main_df], ignore_index=True).to_excel(output_path, index=False)

    # 打开Excel用于写入标注
    wb = load_workbook(output_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 4. 准备主表用于合并
    # 获取主表的合同号列名
    contract_col_main = get_contract_col_name_from_sheet(main_df) # 假设您有这个函数
    
    # 存储原始索引，用于 openpyxl 定位
    main_df['__ROW_IDX__'] = main_df.index
    
    # VVVV 插入归一化函数 VVVV
    # 创建标准合并Key
    main_df['__KEY__'] = normalize_contract_key(main_df[contract_col_main])
    # ^^^^ 插入归一化函数 ^^^^
    
    # 获取本表所有合同号（用于统计等）
    contracts_seen = set(main_df['__KEY__'].dropna())
    # 存储原始索引，用于 openpyxl 定位
    main_df['__ROW_IDX__'] = main_df.index
    # 创建标准合并Key
    main_df['__KEY__'] = main_df[contract_col_main].astype(str).str.strip()
    
    # 获取本表所有合同号
    contracts_seen = set(main_df['__KEY__'].dropna())

    # 5. 一次性合并所有参考数据
    merged_df = main_df.copy()
    for prefix, std_df in ref_dfs_std_dict.items():
        if not std_df.empty:
            merged_df = pd.merge(merged_df, std_df, on='__KEY__', how='left')
    
    total_errors = 0
    skip_city_manager = [0]
    errors_locations = set() # 存储 (row_idx, col_name)
    row_has_error = pd.Series(False, index=merged_df.index) # 标记哪一行有错误

    # 添加Streamlit进度条
    progress = st.progress(0)
    status = st.empty()

    # 6. === 遍历字段进行向量化比对 ===
    mappings_all = {
        'fk': (mapping_fk, ref_dfs_std_dict['fk']),
        'zd': (mapping_zd, ref_dfs_std_dict['zd']),
        'ec': (mapping_ec, ref_dfs_std_dict['ec']),
        'zk': (mapping_zk, ref_dfs_std_dict['zk'])
    }
    
    total_comparisons = sum(len(m[0]) for m in mappings_all.values())
    current_comparison = 0

    for prefix, (mapping, std_df) in mappings_all.items():
        if std_df.empty:
            current_comparison += len(mapping) # 跳过空表
            continue
            
        for main_kw, ref_kw in mapping.items():
            current_comparison += 1
            status.text(f"检查「{sheet_keyword}」: {prefix} - {main_kw}...")
            
            # 关键：在原始 main_df 中找到列名
            exact = (main_kw == "城市经理")
            main_col = find_col(main_df, main_kw, exact=exact)
            
            # 参考列的列名是我们在 prepare_ref_df 中标准化的
            ref_col = f'ref_{prefix}_{main_kw}'

            if not main_col or ref_col not in merged_df.columns:
                continue # 跳过不存在的列

            s_main = merged_df[main_col]
            s_ref = merged_df[ref_col]

            # 处理 "城市经理" 跳过逻辑
            skip_mask = pd.Series(False, index=merged_df.index)
            if main_kw == "城市经理":
                na_strings = ["", "-", "nan", "none", "null"]
                # 检查参考列是否为空
                skip_mask = pd.isna(s_ref) | s_ref.astype(str).str.strip().isin(na_strings)
                skip_city_manager[0] += skip_mask.sum()
            
            # 7. 获取向量化比较结果
            errors_mask = compare_series_vec(s_main, s_ref, main_kw)
            
            # 应用跳过逻辑：如果 skip_mask 为 True，则不算错误
            final_errors_mask = errors_mask & ~skip_mask
            
            if final_errors_mask.any():
                total_errors += final_errors_mask.sum()
                row_has_error |= final_errors_mask
                
                # 8. 存储错误位置 (使用 __ROW_IDX__ 和 原始 main_col 名称)
                bad_indices = merged_df[final_errors_mask]['__ROW_IDX__']
                for idx in bad_indices:
                    errors_locations.add((idx, main_col))
                    
            progress.progress(current_comparison / total_comparisons)

    status.text(f"「{sheet_keyword}」比对完成，正在生成标注文件...")

    # 9. === 遍历错误进行Excel标注 ===
    # (这比遍历所有单元格快得多)
    
    # 获取原始列名 (去掉我们添加的辅助列)
    original_cols_list = list(main_df.drop(columns=['__ROW_IDX__', '__KEY__']).columns)
    # 创建列名到Excel列索引(1-based)的映射
    col_name_to_idx = {name: i + 1 for i, name in enumerate(original_cols_list)}

    # 标红错误单元格
    for (row_idx, col_name) in errors_locations:
        if col_name in col_name_to_idx:
            # +3: (1-based index) + (1 for header) + (1 for empty row)
            ws.cell(row_idx + 3, col_name_to_idx[col_name]).fill = red_fill

    # 标黄有错误的合同号
    if contract_col_main in col_name_to_idx:
        contract_col_excel_idx = col_name_to_idx[contract_col_main]
        # 找到所有出错的原始行号
        error_row_indices = merged_df[row_has_error]['__ROW_IDX__']
        for row_idx in error_row_indices:
            ws.cell(row_idx + 3, contract_col_excel_idx).fill = yellow_fill

    # 10. 导出检查结果
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    st.download_button(
        label=f"📥 下载 {sheet_keyword}审核标注版",
        data=output,
        file_name=f"记录表_{sheet_keyword}_审核标注版.xlsx",
        key=f"download_{sheet_keyword}" # 增加key避免streamlit重跑问题
    )

    elapsed = time.time() - start_time
    st.success(f"✅ {sheet_keyword} 检查完成，共 {total_errors} 处错误，用时 {elapsed:.2f} 秒。")
    return total_errors, elapsed, skip_city_manager[0], contracts_seen

# =====================================
# 📖 文件读取：按关键字识别五份文件
# =====================================
main_file = find_file(uploaded_files, "记录表")
fk_file = find_file(uploaded_files, "放款明细")
zd_file = find_file(uploaded_files, "字段")
ec_file = find_file(uploaded_files, "二次明细")
zk_file = find_file(uploaded_files, "重卡数据")

# 各文件sheet读取（模糊匹配sheet名）
fk_df = pd.read_excel(pd.ExcelFile(fk_file), sheet_name=find_sheet(pd.ExcelFile(fk_file), "本司"))
zd_df = pd.read_excel(pd.ExcelFile(zd_file), sheet_name=find_sheet(pd.ExcelFile(zd_file), "重卡"))
ec_df = pd.read_excel(ec_file)
zk_df = pd.read_excel(zk_file)

# 合同列定位
contract_col_fk = find_col(fk_df, "合同")
contract_col_zd = find_col(zd_df, "合同")
contract_col_ec = find_col(ec_df, "合同")
contract_col_zk = find_col(zk_df, "合同")

# 对照字段映射表
mapping_fk = {"授信方": "授信", "租赁本金": "本金", "租赁期限月": "租赁期限月", "客户经理": "客户经理", "起租收益率": "收益率", "主车台数": "主车台数", "挂车台数": "挂车台数"}
mapping_zd = {"保证金比例": "保证金比例_2", "项目提报人": "提报", "起租时间": "起租日_商", "租赁期限月": "总期数_商_资产", "所属省区": "区域", "城市经理": "城市经理"}
mapping_ec = {"二次时间": "出本流程时间"}
mapping_zk = {"授信方": "授信方"}

# =====================================
# 🚀 (新) 预处理所有参考表
# =====================================
st.info("ℹ️ 正在预处理参考数据...")

# (fk_df, zd_df, ec_df, zk_df 必须已经加载)
fk_std = prepare_ref_df(fk_df, mapping_fk, 'fk')
zd_std = prepare_ref_df(zd_df, mapping_zd, 'zd')
ec_std = prepare_ref_df(ec_df, mapping_ec, 'ec')
zk_std = prepare_ref_df(zk_df, mapping_zk, 'zk')

# 将所有预处理过的DF存入字典，传递给检查函数
ref_dfs_std_dict = {
    'fk': fk_std,
    'zd': zd_std,
    'ec': ec_std,
    'zk': zk_std
}
st.success("✅ 参考数据预处理完成。")

# =====================================
# 🧾 多sheet循环 + 驻店客户表
# =====================================
sheet_keywords = ["二次", "部分担保", "随州", "驻店客户"]
total_all = elapsed_all = skip_total = 0
contracts_seen_all_sheets = set()

# 循环处理四张sheet (调用新函数)
for kw in sheet_keywords:
    # 将 main_file 和 ref_dfs_std_dict 传递进去
    count, used, skipped, seen = check_one_sheet(kw, main_file, ref_dfs_std_dict)
    
    total_all += count
    elapsed_all += used or 0
    skip_total += skipped
    contracts_seen_all_sheets.update(seen)

st.success(f"🎯 全部审核完成，共 {total_all} 处错误，总耗时 {elapsed_all:.2f} 秒。")

# =====================================
# 🕵️ 漏填检查：跳过“是否车管家=是”与“提成类型=联合租赁/驻店”
# =====================================
field_contracts = zd_df[contract_col_zd].dropna().astype(str).str.strip()
col_car_manager = find_col(zd_df, "是否车管家", exact=True)
col_bonus_type = find_col(zd_df, "提成类型", exact=True)

missing_contracts_mask = (~field_contracts.isin(contracts_seen_all_sheets))

# 跳过“车管家=是”
if col_car_manager:
    missing_contracts_mask &= ~(zd_df[col_car_manager].astype(str).str.strip().str.lower() == "是")
# 跳过“联合租赁/驻店”
if col_bonus_type:
    missing_contracts_mask &= ~(
        zd_df[col_bonus_type].astype(str).str.strip().isin(["联合租赁", "驻店"])
    )

# 标记漏填
zd_df_missing = zd_df.copy()
zd_df_missing["漏填检查"] = ""
zd_df_missing.loc[missing_contracts_mask, "漏填检查"] = "❗ 漏填"
漏填合同数 = zd_df_missing["漏填检查"].eq("❗ 漏填").sum()
st.warning(f"⚠️ 共发现 {漏填合同数} 个合同在记录表中未出现（已排除车管家、联合租赁、驻店）")

# =====================================
# 📤 导出字段表（含漏填标注 + 仅漏填版）
# =====================================
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 全字段表（含漏填标注）
wb = Workbook()
ws = wb.active
for c_idx, c in enumerate(zd_df_missing.columns, 1): ws.cell(1, c_idx, c)
for r_idx, row in enumerate(zd_df_missing.itertuples(index=False), 2):
    for c_idx, v in enumerate(row, 1):
        ws.cell(r_idx, c_idx, v)
        if zd_df_missing.columns[c_idx-1] == "漏填检查" and v == "❗ 漏填":
            ws.cell(r_idx, c_idx).fill = yellow_fill

output_all = BytesIO()
wb.save(output_all)
output_all.seek(0)
st.download_button("📥 下载字段表漏填标注版", output_all, "字段表_漏填标注版.xlsx")

# 仅漏填合同
zd_df_only_missing = zd_df_missing[zd_df_missing["漏填检查"] == "❗ 漏填"].copy()
if not zd_df_only_missing.empty:
    wb2 = Workbook()
    ws2 = wb2.active
    for c_idx, c in enumerate(zd_df_only_missing.columns, 1): ws2.cell(1, c_idx, c)
    for r_idx, row in enumerate(zd_df_only_missing.itertuples(index=False), 2):
        for c_idx, v in enumerate(row, 1):
            ws2.cell(r_idx, c_idx, v)
            if zd_df_only_missing.columns[c_idx-1] == "漏填检查" and v == "❗ 漏填":
                ws2.cell(r_idx, c_idx).fill = yellow_fill
    out2 = BytesIO()
    wb2.save(out2)
    out2.seek(0)
    st.download_button("📥 下载仅漏填字段表", out2, "字段表_仅漏填.xlsx")

# =====================================
# ✅ 结束提示
# =====================================
st.success("✅ 所有检查、标注与导出完成！")
